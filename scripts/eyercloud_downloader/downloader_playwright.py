"""
EyerCloud Image Downloader - Versão com Playwright
===================================================
Este script usa o Playwright para automatizar um navegador real,
permitindo capturar cookies HttpOnly e baixar as imagens.

Instalação:
    pip install playwright openpyxl
    playwright install chromium

Uso:
    python downloader_playwright.py                    # Baixa tudo
    python downloader_playwright.py --page 2           # Começa da página 2
    python downloader_playwright.py --report           # Gera apenas o relatório
    python downloader_playwright.py --reset            # Limpa estado e baixa tudo

Na primeira execução, o script vai abrir o navegador e esperar você fazer login.
Depois, ele salva a sessão e nas próximas vezes não precisa logar de novo.
"""

import asyncio
import os
import json
import re
import argparse
from pathlib import Path
from datetime import datetime

# Tenta importar o playwright
try:
    from playwright.async_api import async_playwright
except ImportError:
    print("=" * 60)
    print("ERRO: Playwright não está instalado!")
    print("Execute os seguintes comandos:")
    print("  pip install playwright")
    print("  playwright install chromium")
    print("=" * 60)
    exit(1)

# --- CONFIGURAÇÃO ---
DOWNLOAD_DIR = Path("downloads")
STATE_FILE = Path("download_state.json")
AUTH_STATE_FILE = Path("auth_state.json")
REPORT_FILE = Path("relatorio_downloads.xlsx")
CSV_REPORT_FILE = Path("relatorio_downloads.csv")
BASE_URL = "https://ec2.eyercloud.com"
API_BASE = "https://eyercloud.com/api/v2/eyercloud"


def load_state():
    """Carrega o estado de downloads anteriores."""
    if STATE_FILE.exists():
        with open(STATE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"downloaded_exams": [], "exam_details": {}}


def save_state(state):
    """Salva o estado atual de downloads."""
    with open(STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(state, f, indent=4, ensure_ascii=False)


def count_downloaded_images(exam_folder):
    """Conta quantas imagens foram baixadas para um exame."""
    if not exam_folder.exists():
        return 0
    return len([f for f in exam_folder.iterdir() if f.is_file() and f.suffix.lower() in ['.jpg', '.jpeg', '.png']])


def generate_report(state):
    """Gera um relatório com o status de todos os downloads."""
    print("\n📊 Gerando relatório...")
    
    report_data = []
    
    for exam_id, details in state.get('exam_details', {}).items():
        patient_name = details.get('patient_name', 'Desconhecido')
        expected_count = details.get('expected_images', 0)
        folder_name = details.get('folder_name', '')
        
        # Conta imagens baixadas
        exam_folder = DOWNLOAD_DIR / folder_name
        downloaded_count = count_downloaded_images(exam_folder)
        
        status = "✅ Completo" if downloaded_count >= expected_count else "⚠️ Incompleto"
        if expected_count == 0:
            status = "📭 Sem imagens"
        
        report_data.append({
            'Paciente': patient_name,
            'CPF': details.get('cpf', ''),
            'Nascimento': details.get('birthday', ''),
            'Clínica': details.get('clinic_name', ''),
            'Data Exame': details.get('exam_date', ''),
            'ID Exame': exam_id,
            'Imagens Esperadas': expected_count,
            'Imagens Baixadas': downloaded_count,
            'Status': status,
            'Pasta': folder_name
        })
    
    # Também verifica pastas que existem mas não estão no estado
    for folder in DOWNLOAD_DIR.iterdir():
        if folder.is_dir():
            folder_name = folder.name
            # Verifica se já está no relatório
            if not any(r['Pasta'] == folder_name for r in report_data):
                downloaded_count = count_downloaded_images(folder)
                report_data.append({
                    'Paciente': folder_name.rsplit('_', 1)[0].replace('_', ' '),
                    'CPF': '',
                    'Nascimento': '',
                    'Clínica': '',
                    'Data Exame': '',
                    'ID Exame': folder_name.rsplit('_', 1)[-1] if '_' in folder_name else 'N/A',
                    'Imagens Esperadas': '?',
                    'Imagens Baixadas': downloaded_count,
                    'Status': '❓ Não rastreado',
                    'Pasta': folder_name
                })
    
    # Ordena por nome do paciente
    report_data.sort(key=lambda x: x['Paciente'])
    
    # Calcula totais
    total_expected = sum(r['Imagens Esperadas'] for r in report_data if isinstance(r['Imagens Esperadas'], int))
    total_downloaded = sum(r['Imagens Baixadas'] for r in report_data)
    
    # Gera CSV (sempre funciona)
    with open(CSV_REPORT_FILE, 'w', encoding='utf-8-sig') as f:
        headers = ['Paciente', 'CPF', 'Nascimento', 'Clínica', 'Data Exame', 'ID Exame', 'Imagens Esperadas', 'Imagens Baixadas', 'Status', 'Pasta']
        f.write(';'.join(headers) + '\n')
        for row in report_data:
            f.write(';'.join(str(row.get(h, '')) for h in headers) + '\n')
        f.write('\n')
        f.write(f'TOTAL;;;;;;{total_expected};{total_downloaded};;\n')
    
    print(f"   📄 CSV salvo em: {CSV_REPORT_FILE.absolute()}")
    
    # Tenta gerar Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Downloads"
        
        # Cabeçalho
        headers = ['Paciente', 'CPF', 'Nascimento', 'Clínica', 'Data Exame', 'ID Exame', 'Imagens Esperadas', 'Imagens Baixadas', 'Status', 'Pasta']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Dados
        for row_idx, row_data in enumerate(report_data, 2):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col, value=row_data.get(header, ''))
                if header == 'Status':
                    if '✅' in str(row_data[header]):
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif '⚠️' in str(row_data[header]):
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Linha de totais
        total_row = len(report_data) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=total_expected).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=total_downloaded).font = Font(bold=True)
        
        # Ajusta largura das colunas
        column_widths = {
            'A': 40, 'B': 15, 'C': 15, 'D': 25, 'E': 25, 
            'F': 15, 'G': 18, 'H': 18, 'I': 15, 'J': 40
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        wb.save(REPORT_FILE)
        print(f"   📊 Excel salvo em: {REPORT_FILE.absolute()}")
        
    except ImportError:
        print("   ℹ️ openpyxl não instalado. Use 'pip install openpyxl' para gerar Excel.")
    
    # Imprime resumo no console
    print("\n" + "=" * 80)
    print(f"{'PACIENTE':<40} {'ESPERADAS':<12} {'BAIXADAS':<12} {'STATUS':<15}")
    print("=" * 80)
    for row in report_data:
        print(f"{row['Paciente'][:38]:<40} {str(row['Imagens Esperadas']):<12} {row['Imagens Baixadas']:<12} {row['Status']:<15}")
    print("=" * 80)
    print(f"{'TOTAL':<40} {total_expected:<12} {total_downloaded:<12}")
    print("=" * 80)
    
    return report_data


async def fetch_exams_via_api(page, current_page=1):
    """Busca a lista de exames via API."""
    payload = {
        "filter": {
            "startDate": None,
            "endDate": None,
            "patientID": None,
            "patientFullName": None,
            "properties": {
                "mcRas": False, "color": False, "redfree": False,
                "infrared": False, "segAnterior": False,
                "panoramic": False, "stereo": False
            }
        },
        "examCurrentPage": current_page
    }
    
    result = await page.evaluate('''async (payload) => {
        let response = await fetch("https://eyercloud.com/api/v2/eyercloud/exam/list", {
            method: "POST",
            headers: {
                "Content-Type": "application/json",
                "Accept": "application/json"
            },
            body: JSON.stringify(payload),
            credentials: "include"
        });
        
        let data = await response.json();
        
        if (!data.result || data.result.length === 0) {
            response = await fetch("https://eyercloud.com/api/v2/eyercloud/exam/filter-20-last-with-examdata-and-params", {
                method: "GET",
                headers: { "Accept": "application/json" },
                credentials: "include"
            });
            
            const text = await response.text();
            try {
                data = JSON.parse(text);
                if (Array.isArray(data)) {
                    data = { result: data };
                } else if (data.exams) {
                    data = { result: data.exams };
                }
            } catch (e) {
                data = { result: [] };
            }
        }
        
        return data;
    }''', payload)
    
    if current_page == 1:
        print(f"    [DEBUG] API retornou: {type(result)} - keys: {list(result.keys()) if isinstance(result, dict) else 'N/A'}")
        if isinstance(result, dict) and 'result' in result:
            print(f"    [DEBUG] Número de exames: {len(result.get('result', []))}")
    
    return result


async def fetch_exam_details(page, exam_id):
    """Busca os detalhes de um exame específico."""
    result = await page.evaluate('''async (examId) => {
        const response = await fetch(`https://eyercloud.com/api/v2/eyercloud/examData/list?id=${examId}`, {
            method: "GET",
            headers: {
                "Accept": "application/json"
            },
            credentials: "include"
        });
        
        const text = await response.text();
        try {
            return JSON.parse(text);
        } catch (e) {
            console.log("Erro ao parsear detalhes do exame:", text.substring(0, 200));
            return { examDataList: [] };
        }
    }''', exam_id)
    
    return result


async def download_image(page, url, filepath):
    """Baixa uma imagem usando o contexto do navegador."""
    if filepath.exists():
        return False
    
    try:
        response = await page.evaluate('''async (url) => {
            const response = await fetch(url);
            if (!response.ok) {
                return { error: response.status };
            }
            const blob = await response.blob();
            const reader = new FileReader();
            return new Promise((resolve) => {
                reader.onloadend = () => resolve({ data: reader.result });
                reader.readAsDataURL(blob);
            });
        }''', url)
        
        if 'error' in response:
            print(f"    ❌ Erro {response['error']} ao baixar: {url}")
            return False
        
        data_url = response['data']
        base64_data = data_url.split(',', 1)[1]
        
        import base64
        image_data = base64.b64decode(base64_data)
        
        filepath.parent.mkdir(parents=True, exist_ok=True)
        with open(filepath, 'wb') as f:
            f.write(image_data)
        
        return True
        
    except Exception as e:
        print(f"    ❌ Exceção ao baixar {url}: {e}")
        return False


async def main():
    parser = argparse.ArgumentParser(description='EyerCloud Image Downloader')
    parser.add_argument('--page', type=int, default=1, help='Página inicial (padrão: 1)')
    parser.add_argument('--report', action='store_true', help='Apenas gera o relatório, sem baixar')
    parser.add_argument('--reset', action='store_true', help='Limpa o estado e baixa tudo novamente')
    parser.add_argument('--retry', action='store_true', help='Tenta baixar novamente apenas os exames incompletos')
    parser.add_argument('--interactive', action='store_true', help='Modo interativo: navegue manualmente e pressione ENTER para baixar')
    args = parser.parse_args()
    
    print("=" * 60)
    print("🏥 EyerCloud Image Downloader (Playwright Edition)")
    print("=" * 60)
    
    state = load_state()
    
    if args.reset:
        print("🔄 Resetando estado...")
        state = {"downloaded_exams": [], "exam_details": {}}
        save_state(state)
    
    if args.retry:
        print("🔄 Modo RETRY: Verificando exames incompletos...")
        # Remove dos "já baixados" os exames que estão incompletos
        incomplete_exams = []
        for exam_id, details in state.get('exam_details', {}).items():
            expected = details.get('expected_images', 0)
            folder_name = details.get('folder_name', '')
            downloaded = count_downloaded_images(DOWNLOAD_DIR / folder_name)
            if downloaded < expected:
                incomplete_exams.append(exam_id)
                print(f"   ⚠️ {details.get('patient_name', 'N/A')}: {downloaded}/{expected}")
        
        # Remove os incompletos da lista de "já baixados"
        state['downloaded_exams'] = [e for e in state['downloaded_exams'] if e not in incomplete_exams]
        save_state(state)
        print(f"   📋 {len(incomplete_exams)} exames serão re-tentados")
    
    if args.report:
        generate_report(state)
        return
    
    DOWNLOAD_DIR.mkdir(exist_ok=True)
    
    async with async_playwright() as p:
        if AUTH_STATE_FILE.exists():
            print("📂 Sessão anterior encontrada. Tentando reutilizar...")
            browser = await p.chromium.launch(headless=False)
            context = await browser.new_context(storage_state=str(AUTH_STATE_FILE))
        else:
            print("🌐 Abrindo navegador para login...")
            browser = await p.chromium.launch(headless=False)
            context = await browser.new_context()
        
        page = await context.new_page()
        
        await page.goto(f"{BASE_URL}/exam", wait_until="networkidle")
        await asyncio.sleep(3)
        
        page_content = await page.content()
        is_logged_in = "Acessar exame" in page_content or "exam-card" in page_content or "patient" in page_content.lower()
        
        if not is_logged_in or "login" in page.url.lower():
            print("\n" + "=" * 60)
            print("🔐 FAÇA LOGIN NO NAVEGADOR QUE ABRIU")
            print("   1. Complete o login com seu email e senha")
            print("   2. Aguarde a lista de exames carregar")
            print("   3. Volte aqui e pressione ENTER para continuar")
            print("=" * 60)
            
            await asyncio.get_event_loop().run_in_executor(None, input, "\n>>> Pressione ENTER após fazer login... ")
            
            await context.storage_state(path=str(AUTH_STATE_FILE))
            print("💾 Sessão salva para uso futuro!")
        else:
            print("✅ Já está logado!")
        
        await page.goto(f"{BASE_URL}/exam", wait_until="networkidle")
        await asyncio.sleep(3)
        
        # MODO INTERATIVO
        if args.interactive:
            print("\n" + "=" * 60)
            print("🎮 MODO INTERATIVO ATIVADO")
            print("=" * 60)
            print("   O navegador vai permanecer aberto.")
            print("   1. Navegue para a página que deseja baixar")
            print("   2. Volte aqui e pressione ENTER para baixar")
            print("   3. Digite 'sair' para encerrar")
            print("=" * 60)
            
            total_downloaded = 0
            
            while True:
                user_input = await asyncio.get_event_loop().run_in_executor(
                    None, input, "\n>>> Pressione ENTER para baixar a página atual (ou 'sair'): "
                )
                
                if user_input.lower().strip() in ['sair', 'exit', 'quit', 'q']:
                    print("👋 Encerrando...")
                    break
                
                print("\n📋 Extraindo exames da página atual (DOM)...")
                
                # Extrai os IDs dos exames diretamente do DOM da página
                try:
                    # Captura os IDs e nomes dos exames visíveis (.patient-box com data-exam-id)
                    exam_data = await page.evaluate('''() => {
                        const exams = [];
                        const patientBoxes = document.querySelectorAll('.patient-box[data-exam-id]');
                        patientBoxes.forEach(box => {
                            const examId = box.getAttribute('data-exam-id');
                            const nameEl = box.querySelector('a.btn-secondary span');
                            const patientName = nameEl ? nameEl.innerText : 'Desconhecido';
                            if (examId) {
                                exams.push({id: examId, name: patientName});
                            }
                        });
                        return exams;
                    }''')
                    
                    if not exam_data:
                        print("❌ Nenhum exame encontrado nesta página.")
                        print("   Dica: Certifique-se de que a lista de exames está visível.")
                        continue
                    
                    print(f"   Encontrados {len(exam_data)} exames na tela")
                    
                    for exam in exam_data:
                        exam_id = exam['id']
                        patient_name = exam['name']
                        
                        if exam_id in state['downloaded_exams']:
                            print(f"  ⏭️ Já baixado: {exam_id[:8]}...")
                            continue
                        
                        # Busca detalhes do exame via API
                        details = await fetch_exam_details(page, exam_id)
                        
                        # Usa o nome do paciente do DOM (mais confiável)
                        safe_name = re.sub(r'[<>:"/\\|?*]', '_', patient_name)
                        safe_name = safe_name.replace(' ', '_')
                        
                        print(f"\n  👤 {patient_name} ({exam_id[:8]}...)")
                        
                        image_list = details.get('examDataList', [])
                        expected_count = len(image_list)
                        
                        folder_name = f"{safe_name}_{exam_id[:8]}"
                        
                        # Extrai dados do paciente do exame
                        exam_obj = details.get('exam', {})
                        patient_obj = exam_obj.get('patient', {})
                        anamnesis = patient_obj.get('anamnesis', {}) or {}
                        
                        # Extrai doenças de base
                        underlying_diseases = {
                            "diabetes": anamnesis.get('diabetes', patient_obj.get('diabetes', False)),
                            "hypertension": anamnesis.get('hypertension', patient_obj.get('hypertension', False)),
                            "cholesterol": anamnesis.get('cholesterol', patient_obj.get('cholesterol', False)),
                            "smoker": anamnesis.get('smoker', patient_obj.get('smoker', False))
                        }
                        
                        # Extrai doenças oftalmológicas
                        ophthalmic_diseases = {
                            "diabeticRetinopathy": anamnesis.get('diabeticRetinopathy', patient_obj.get('diabeticRetinopathy', False)),
                            "dmri": anamnesis.get('dmri', patient_obj.get('dmri', False)),
                            "glaucoma": anamnesis.get('glaucoma', patient_obj.get('glaucoma', False)),
                            "cataract": anamnesis.get('cataract', patient_obj.get('cataract', False)),
                            "pterygium": anamnesis.get('pterygium', patient_obj.get('pterygium', False)),
                            "lowVisualAcuity": anamnesis.get('lowVisualAcuity', patient_obj.get('lowVisualAcuity', False))
                        }
                        
                        state['exam_details'][exam_id] = {
                            'patient_name': patient_name,
                            'expected_images': expected_count,
                            'folder_name': folder_name,
                            'exam_date': exam_obj.get('date'),
                            'cpf': patient_obj.get('cpf', ''),
                            'birthday': patient_obj.get('birthday', ''),
                            'gender': patient_obj.get('gender', ''),
                            'clinic_name': exam_obj.get('clinic', {}).get('name', '') or patient_obj.get('place', ''),
                            'underlying_diseases': underlying_diseases,
                            'ophthalmic_diseases': ophthalmic_diseases,
                            'otherDisease': patient_obj.get('otherDisease') or anamnesis.get('otherDisease'),
                            'download_date': datetime.now().isoformat()
                        }
                        save_state(state)
                        
                        if not image_list:
                            print("    📭 Sem imagens neste exame.")
                            state['downloaded_exams'].append(exam_id)
                            save_state(state)
                            continue
                        
                        data_path = details.get('dataPath', 'https://d25chn8x2vrs37.cloudfront.net')
                        exam_folder = DOWNLOAD_DIR / folder_name
                        
                        exam_downloaded = 0
                        for img_data in image_list:
                            uuid = img_data['uuid']
                            img_url = f"{data_path}/{uuid}"
                            filepath = exam_folder / f"{uuid}.jpg"
                            
                            if await download_image(page, img_url, filepath):
                                print(f"    ✅ {uuid[:8]}...")
                                exam_downloaded += 1
                                total_downloaded += 1
                        
                        print(f"    📸 {exam_downloaded}/{expected_count} imagens baixadas")
                        state['downloaded_exams'].append(exam_id)
                        save_state(state)
                        
                        await asyncio.sleep(0.5)
                    
                except Exception as e:
                    print(f"❌ Erro: {e}")
                    import traceback
                    traceback.print_exc()
            
            # Fim do modo interativo
            await browser.close()
            generate_report(state)
            return
        
        # MODO AUTOMÁTICO (código original)
        print(f"\n📋 Buscando lista de exames (começando da página {args.page})...")
        
        page_num = args.page
        total_downloaded = 0
        all_seen_ids = set()
        
        while True:
            print(f"\n📄 Página {page_num}...")
            
            try:
                data = await fetch_exams_via_api(page, page_num)
                exams = data.get('result', [])
                
                if not exams:
                    print("✅ Não há mais exames.")
                    break
                
                current_ids = {e['id'] for e in exams}
                if current_ids.issubset(all_seen_ids):
                    print("✅ Fim da lista (API repetindo resultados).")
                    break
                all_seen_ids.update(current_ids)
                
                for exam in exams:
                    exam_id = exam['id']
                    
                    if exam_id in state['downloaded_exams']:
                        print(f"  ⏭️ Já baixado: {exam_id[:8]}...")
                        continue
                    
                    patient_name = (
                        exam.get('patientFullName') or 
                        exam.get('patientName') or 
                        exam.get('name') or
                        'Paciente_Desconhecido'
                    )
                    if isinstance(exam.get('patient'), dict):
                        patient_name = exam['patient'].get('fullName') or exam['patient'].get('name') or patient_name
                    
                    safe_name = re.sub(r'[<>:"/\\|?*]', '_', patient_name)
                    safe_name = safe_name.replace(' ', '_')
                    
                    print(f"\n  👤 {patient_name} ({exam_id[:8]}...)")
                    
                    details = await fetch_exam_details(page, exam_id)
                    
                    image_list = details.get('examDataList', [])
                    expected_count = len(image_list)
                    
                    folder_name = f"{safe_name}_{exam_id[:8]}"
                    
                    # Extrai dados do paciente do exame
                    exam_obj = details.get('exam', {})
                    patient_obj = exam_obj.get('patient', {})
                    anamnesis = patient_obj.get('anamnesis', {}) or {}
                    
                    # Também tenta pegar dados do objeto exam original da lista
                    clinic_name = exam_obj.get('clinic', {}).get('name', '')
                    if not clinic_name:
                        clinic_name = exam.get('clinicName', '') or exam.get('clinic', {}).get('name', '')
                    
                    # Extrai doenças de base
                    underlying_diseases = {
                        "diabetes": anamnesis.get('diabetes', patient_obj.get('diabetes', False)),
                        "hypertension": anamnesis.get('hypertension', patient_obj.get('hypertension', False)),
                        "cholesterol": anamnesis.get('cholesterol', patient_obj.get('cholesterol', False)),
                        "smoker": anamnesis.get('smoker', patient_obj.get('smoker', False))
                    }
                    
                    # Extrai doenças oftalmológicas
                    ophthalmic_diseases = {
                        "diabeticRetinopathy": anamnesis.get('diabeticRetinopathy', patient_obj.get('diabeticRetinopathy', False)),
                        "dmri": anamnesis.get('dmri', patient_obj.get('dmri', False)),
                        "glaucoma": anamnesis.get('glaucoma', patient_obj.get('glaucoma', False)),
                        "cataract": anamnesis.get('cataract', patient_obj.get('cataract', False)),
                        "pterygium": anamnesis.get('pterygium', patient_obj.get('pterygium', False)),
                        "lowVisualAcuity": anamnesis.get('lowVisualAcuity', patient_obj.get('lowVisualAcuity', False))
                    }
                    
                    # Salva detalhes para o relatório
                    state['exam_details'][exam_id] = {
                        'patient_name': patient_name,
                        'expected_images': expected_count,
                        'folder_name': folder_name,
                        'exam_date': exam_obj.get('date'),
                        'cpf': patient_obj.get('cpf', '') or exam.get('patient', {}).get('cpf', ''),
                        'birthday': patient_obj.get('birthday', '') or exam.get('patient', {}).get('birthday', ''),
                        'gender': patient_obj.get('gender', '') or exam.get('patient', {}).get('gender', ''),
                        'clinic_name': clinic_name,
                        'underlying_diseases': underlying_diseases,
                        'ophthalmic_diseases': ophthalmic_diseases,
                        'otherDisease': patient_obj.get('otherDisease') or anamnesis.get('otherDisease'),
                        'download_date': datetime.now().isoformat()
                    }
                    save_state(state)
                    
                    if not image_list:
                        print("    📭 Sem imagens neste exame.")
                        state['downloaded_exams'].append(exam_id)
                        save_state(state)
                        continue
                    
                    data_path = details.get('dataPath', 'https://d25chn8x2vrs37.cloudfront.net')
                    exam_folder = DOWNLOAD_DIR / folder_name
                    
                    exam_downloaded = 0
                    for img_data in image_list:
                        uuid = img_data['uuid']
                        img_url = f"{data_path}/{uuid}"
                        filepath = exam_folder / f"{uuid}.jpg"
                        
                        if await download_image(page, img_url, filepath):
                            print(f"    ✅ {uuid[:8]}...")
                            exam_downloaded += 1
                            total_downloaded += 1
                    
                    print(f"    📸 {exam_downloaded}/{expected_count} imagens baixadas")
                    state['downloaded_exams'].append(exam_id)
                    save_state(state)
                    
                    await asyncio.sleep(0.5)
                
                page_num += 1
                await asyncio.sleep(1)
                
            except Exception as e:
                print(f"❌ Erro: {e}")
                import traceback
                traceback.print_exc()
                break
        
        print("\n" + "=" * 60)
        print(f"🎉 Concluído! Total de imagens baixadas nesta sessão: {total_downloaded}")
        print(f"📁 Pasta de downloads: {DOWNLOAD_DIR.absolute()}")
        print("=" * 60)
        
        await browser.close()
    
    # Gera relatório automaticamente ao final
    generate_report(state)


if __name__ == "__main__":
    asyncio.run(main())
