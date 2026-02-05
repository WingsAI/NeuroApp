#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Regenera o relatório CSV/XLSX usando os dados do bytescale_mapping.json
que contém as informações completas dos pacientes.
"""

import json
from pathlib import Path
from datetime import datetime

# Arquivos
BYTESCALE_MAPPING = Path("bytescale_mapping_cleaned.json")
CSV_REPORT_FILE = Path("relatorio_downloads.csv")
XLSX_REPORT_FILE = Path("relatorio_downloads.xlsx")
DOWNLOAD_DIR = Path("downloads")


def count_downloaded_images(exam_folder):
    """Conta quantas imagens foram baixadas para um exame."""
    if not exam_folder.exists():
        return 0
    return len([f for f in exam_folder.iterdir() if f.is_file() and f.suffix.lower() in ['.jpg', '.jpeg', '.png']])


def format_cpf(cpf):
    """Formata CPF para exibição."""
    if not cpf or len(cpf) != 11:
        return cpf or ''
    return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"


def format_birthday(birthday):
    """Formata data de nascimento."""
    if not birthday:
        return ''
    try:
        # Tenta parsear varios formatos
        if 'T' in birthday:
            dt = datetime.fromisoformat(birthday.replace('Z', '+00:00'))
            return dt.strftime('%d/%m/%Y')
        return birthday
    except:
        return birthday


def main():
    if not BYTESCALE_MAPPING.exists():
        print(f"❌ Arquivo {BYTESCALE_MAPPING} não encontrado.")
        print("   Execute primeiro o bytescale_uploader.py")
        return
    
    print("📊 Regenerando relatório a partir do bytescale_mapping.json...")
    
    with open(BYTESCALE_MAPPING, 'r', encoding='utf-8') as f:
        mapping = json.load(f)
    
    print(f"   Encontrados {len(mapping)} pacientes no mapping")
    
    report_data = []
    
    for patient_key, patient_data in mapping.items():
        patient_name = patient_data.get('patient_name', 'Desconhecido')
        exam_id = patient_data.get('exam_id', '')
        folder_name = patient_key
        
        # Conta imagens
        images = patient_data.get('images', [])
        expected_count = len(images)
        
        exam_folder = DOWNLOAD_DIR / folder_name
        downloaded_count = count_downloaded_images(exam_folder)
        
        # Determina status
        if expected_count == 0:
            status = "📭 Sem imagens"
        elif downloaded_count >= expected_count:
            status = "✅ Completo"
        else:
            status = "⚠️ Incompleto"
        
        # Extrai upload date da primeira imagem (se houver)
        upload_date = ''
        if images:
            upload_date = images[0].get('upload_date', '')
        
        report_data.append({
            'Paciente': patient_name,
            'CPF': format_cpf(patient_data.get('cpf', '')),
            'Nascimento': format_birthday(patient_data.get('birthday', '')),
            'Clínica': patient_data.get('clinic_name', '').strip(),
            'Data Exame': '',  # Não temos essa info diretamente
            'ID Exame': exam_id,
            'Imagens Esperadas': expected_count,
            'Imagens Baixadas': downloaded_count,
            'Status': status,
            'Pasta': folder_name
        })
    
    # Ordena por nome
    report_data.sort(key=lambda x: x['Paciente'])
    
    # Calcula totais
    total_expected = sum(r['Imagens Esperadas'] for r in report_data if isinstance(r['Imagens Esperadas'], int))
    total_downloaded = sum(r['Imagens Baixadas'] for r in report_data)
    
    # Gera CSV
    with open(CSV_REPORT_FILE, 'w', encoding='utf-8-sig') as f:
        headers = ['Paciente', 'CPF', 'Nascimento', 'Clínica', 'Data Exame', 'ID Exame', 'Imagens Esperadas', 'Imagens Baixadas', 'Status', 'Pasta']
        f.write(';'.join(headers) + '\n')
        for row in report_data:
            f.write(';'.join(str(row.get(h, '')) for h in headers) + '\n')
        f.write('\n')
        f.write(f'TOTAL;;;;;;{total_expected};{total_downloaded};;\n')
    
    print(f"   📄 CSV salvo em: {CSV_REPORT_FILE.absolute()}")
    
    # Tenta gerar Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Downloads"
        
        headers = ['Paciente', 'CPF', 'Nascimento', 'Clínica', 'Data Exame', 'ID Exame', 'Imagens Esperadas', 'Imagens Baixadas', 'Status', 'Pasta']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, row_data in enumerate(report_data, 2):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col, value=row_data.get(header, ''))
                if header == 'Status':
                    if '✅' in str(row_data[header]):
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif '⚠️' in str(row_data[header]):
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        total_row = len(report_data) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=total_expected).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=total_downloaded).font = Font(bold=True)
        
        column_widths = {
            'A': 40, 'B': 18, 'C': 12, 'D': 25, 'E': 20,
            'F': 30, 'G': 18, 'H': 18, 'I': 20, 'J': 60
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        wb.save(XLSX_REPORT_FILE)
        print(f"   📊 Excel salvo em: {XLSX_REPORT_FILE.absolute()}")
    
    except ImportError:
        print("   ⚠️ openpyxl não instalado. Execute: pip install openpyxl")
    except Exception as e:
        print(f"   ⚠️ Erro ao gerar Excel: {e}")
    
    print(f"\n✅ Relatório gerado com {len(report_data)} pacientes")
    print(f"   Total esperado: {total_expected} imagens")
    print(f"   Total baixado:  {total_downloaded} imagens")


if __name__ == "__main__":
    main()
