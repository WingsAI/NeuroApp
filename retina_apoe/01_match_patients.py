"""
01_match_patients.py — Cruzamento de pacientes entre laudos CSV e planilha APOE Tauá.

Gera matched_patients.csv com dados unificados para análise.

Uso:
    python 01_match_patients.py              # Preview: mostra matches e stats
    python 01_match_patients.py --execute    # Gera matched_patients.csv
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import csv
import json
import os
import re
import unicodedata
from pathlib import Path
from difflib import SequenceMatcher

SCRIPT_DIR = Path(__file__).parent
CSV_PATH = SCRIPT_DIR / 'laudados_2026-02-26.csv'
DOWNLOADS_DIR = SCRIPT_DIR.parent / 'downloads'

# Manual exclusions: fuzzy matches that are clearly wrong people
FALSE_MATCHES = {
    # CSV name -> APOE name (normalized) — wrong person matches
    'FRANCISDALVA ALVES DA SILVA': True,         # ≠ MARINALVA ALVES DA SILVA
    'MARIA DA CONCEICAO SOUSA E SILVA': True,    # ≠ MARIA DA CONCEICAO SOUSA DOS SANTOS
    'LEONOR DO NASCIMENTO SILVA': True,           # ≠ LEONARDO NASCIMENTO SILVA (diff gender)
    'CRISTINA CANDIDO DA FRANCA': True,           # ≠ CRISTIANE CANDIDO DA FRANCA
    'FRANCISCO CAMILO DE SOUSA': True,            # ≠ FRANCISCO CAMILO DE SANTOS
    'MARIA DAS GRACAS PEREIRA LIMA': True,        # ≠ MARIA DAS GRACAS PEREIRA LO (truncated?)
}

# --- Normalization ---

def normalize_name(name):
    """Normalize name: uppercase, strip accents, collapse spaces."""
    if not name:
        return ''
    name = unicodedata.normalize('NFKD', name)
    name = ''.join(c for c in name if not unicodedata.combining(c))
    name = name.upper().strip()
    name = re.sub(r'\s+', ' ', name)
    return name


def name_similarity(a, b):
    """Compute similarity between two names using multiple strategies."""
    if not a or not b:
        return 0.0

    # Exact match
    if a == b:
        return 1.0

    # SequenceMatcher ratio
    seq_ratio = SequenceMatcher(None, a, b).ratio()

    # Word overlap score
    words_a = set(a.split())
    words_b = set(b.split())
    if not words_a or not words_b:
        return seq_ratio
    common = words_a & words_b
    # Jaccard-like but weighted toward the shorter name
    word_ratio = len(common) / min(len(words_a), len(words_b))

    # Handle abbreviations: "R." matching "RODRIGUES", "P." matching "PEREIRA"
    abbrev_bonus = 0
    for wa in words_a:
        if len(wa) <= 2 and wa.endswith('.'):
            letter = wa[0]
            for wb in words_b:
                if wb.startswith(letter) and len(wb) > 2:
                    abbrev_bonus += 0.05
    for wb in words_b:
        if len(wb) <= 2 and wb.endswith('.'):
            letter = wb[0]
            for wa in words_a:
                if wa.startswith(letter) and len(wa) > 2:
                    abbrev_bonus += 0.05

    return max(seq_ratio, word_ratio) + min(abbrev_bonus, 0.1)


# --- Load APOE data ---

def load_apoe_taua():
    """Load APOE Tauá Excel. Returns list of dicts."""
    import openpyxl

    fname = None
    for f in os.listdir(SCRIPT_DIR):
        if f.startswith('APOE') and f.endswith('.xlsx'):
            fname = SCRIPT_DIR / f
            break
    if not fname:
        raise FileNotFoundError("APOE_TAUÁ_ID.xlsx not found")

    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    patients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        name = data.get('Nome completo')
        if not name:
            continue

        # Determine genotype from ε columns
        genotype = None
        for g in ['ε2ε2', 'ε2ε3', 'ε2ε4', 'ε3ε3', 'ε3ε4', 'ε4ε4']:
            if data.get(g) == 1:
                genotype = g.replace('ε', 'e')
                break

        # If genotype not from binary columns, try from e2/e3/e4 allele markers
        if not genotype:
            alleles = []
            if data.get('e2') and str(data['e2']).strip().lower() == 'x':
                alleles.append('e2')
            if data.get('e3') and str(data['e3']).strip().lower() == 'x':
                alleles.append('e3')
            if data.get('e4') and str(data['e4']).strip().lower() == 'x':
                alleles.append('e4')
            if len(alleles) == 2:
                genotype = alleles[0] + alleles[1]
            elif len(alleles) == 1:
                genotype = alleles[0] + alleles[0]

        if not genotype:
            # Check 'Repetir' column — might need to redo genotyping
            if data.get('Repetir'):
                continue  # Skip patients needing retest
            continue

        patients.append({
            'name': str(name).strip(),
            'name_normalized': normalize_name(str(name)),
            'genotype': genotype,
            'age': data.get('Idade'),
            'sex': data.get('Sexo'),  # F or M
            'has': str(data.get('HAS', '')).upper().strip() if data.get('HAS') else None,
            'dm': str(data.get('DM', '')).upper().strip() if data.get('DM') else None,
            'dislipidemia': str(data.get('DISLIPIDEMIA', '')).upper().strip() if data.get('DISLIPIDEMIA') else None,
            'escolaridade': data.get('Escolaridade'),
            'cor_raca': data.get('  Cor/raça'),  # note leading spaces in header
        })

    return patients


# --- Load CSV laudos ---

def load_laudos_taua():
    """Load laudos CSV, filter Tauá only. Returns list of dicts."""
    patients = []

    with open(CSV_PATH, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            location = row.get('Unidade/Local', '')
            if 'Tauá' not in location:
                continue

            name = row.get('Nome', '').strip()
            if not name:
                continue

            # Parse achados JSON for quality
            achados = row.get('Achados', '')
            od_quality = 'unknown'
            oe_quality = 'unknown'
            try:
                a = json.loads(achados)
                od_quality = a.get('od', {}).get('quality', 'unknown')
                oe_quality = a.get('oe', {}).get('quality', 'unknown')
            except (json.JSONDecodeError, TypeError):
                pass

            url_od = row.get('Imagem OD (URL)', '').strip()
            url_oe = row.get('Imagem OE (URL)', '').strip()

            patients.append({
                'name': name,
                'name_normalized': normalize_name(name),
                'age': row.get('Idade', ''),
                'sex': row.get('Gênero', ''),
                'cpf': row.get('CPF', ''),
                'birth_date': row.get('Data Nascimento', ''),
                'exam_date': row.get('Data Exame', ''),
                'od_quality': od_quality,
                'oe_quality': oe_quality,
                'url_od': url_od if od_quality != 'impossible' else '',
                'url_oe': url_oe if oe_quality != 'impossible' else '',
                'has_db': row.get('DB: Hipertensão', ''),
                'dm_db': row.get('DB: Diabetes', ''),
                'diagnosis': row.get('Diagnóstico', ''),
            })

    return patients


def url_to_local_path(url):
    """Convert Bytescale URL to local file path in downloads/."""
    if not url:
        return ''
    # URL: https://upcdn.io/W142icY/raw/neuroapp/patients/PATIENT_NAME_EXAMID/UUID.jpg
    # Local: downloads/PATIENT_NAME_EXAMID/UUID.jpg
    match = re.search(r'/patients/([^/]+)/([^/]+\.jpg)', url)
    if not match:
        return ''
    folder = match.group(1)
    filename = match.group(2)
    local = DOWNLOADS_DIR / folder / filename
    if local.exists():
        return str(local)
    return ''


# --- Matching ---

def match_patients(laudos, apoe_patients):
    """Match laudos patients to APOE patients by normalized name with fuzzy matching."""
    matches = []
    unmatched_csv = []
    used_apoe = set()

    # Build lookup for exact matches
    apoe_by_name = {}
    for ap in apoe_patients:
        apoe_by_name.setdefault(ap['name_normalized'], []).append(ap)

    for laudo in laudos:
        ln = laudo['name_normalized']

        # 1. Try exact match
        if ln in apoe_by_name:
            ap = apoe_by_name[ln][0]
            matches.append((laudo, ap, 1.0, 'exact'))
            used_apoe.add(ap['name_normalized'])
            continue

        # 2. Fuzzy match
        best_score = 0
        best_ap = None
        for ap in apoe_patients:
            if ap['name_normalized'] in used_apoe:
                continue
            score = name_similarity(ln, ap['name_normalized'])
            if score > best_score:
                best_score = score
                best_ap = ap

        if best_score >= 0.80 and best_ap and ln not in FALSE_MATCHES:
            matches.append((laudo, best_ap, best_score, 'fuzzy'))
            used_apoe.add(best_ap['name_normalized'])
        else:
            unmatched_csv.append((laudo, best_ap, best_score))

    return matches, unmatched_csv


def main():
    execute = '--execute' in sys.argv

    print("=" * 70)
    print("  RETINA x APOE — Cruzamento de Pacientes")
    print("=" * 70)
    print()

    # Load data
    print("Carregando dados...")
    apoe_patients = load_apoe_taua()
    laudos = load_laudos_taua()
    print(f"  APOE Tauá: {len(apoe_patients)} pacientes com genótipo")
    print(f"  Laudos Tauá: {len(laudos)} pacientes")
    print()

    # Match
    print("Cruzando nomes...")
    matches, unmatched = match_patients(laudos, apoe_patients)
    print(f"  Matched: {len(matches)}")
    print(f"  Unmatched: {len(unmatched)}")
    print()

    # Analyze matches
    exact = [m for m in matches if m[3] == 'exact']
    fuzzy = [m for m in matches if m[3] == 'fuzzy']
    print(f"  Exact matches: {len(exact)}")
    print(f"  Fuzzy matches: {len(fuzzy)}")
    print()

    # Show fuzzy matches for review
    if fuzzy:
        print("--- FUZZY MATCHES (review) ---")
        for laudo, ap, score, _ in sorted(fuzzy, key=lambda x: x[2]):
            print(f"  [{score:.2f}] CSV: {laudo['name_normalized']}")
            print(f"         APOE: {ap['name_normalized']}")
            print()

    # Image availability
    total_images = 0
    patients_with_images = 0
    results = []

    for laudo, ap, score, match_type in matches:
        local_od = url_to_local_path(laudo['url_od'])
        local_oe = url_to_local_path(laudo['url_oe'])

        has_od = bool(local_od) and laudo['od_quality'] != 'impossible'
        has_oe = bool(local_oe) and laudo['oe_quality'] != 'impossible'

        if has_od or has_oe:
            patients_with_images += 1
            if has_od: total_images += 1
            if has_oe: total_images += 1

        results.append({
            'name_csv': laudo['name'],
            'name_apoe': ap['name'],
            'match_score': f"{score:.2f}",
            'match_type': match_type,
            'genotype': ap['genotype'],
            'age_apoe': ap.get('age', ''),
            'sex_apoe': ap.get('sex', ''),
            'has_apoe': ap.get('has', ''),
            'dm_apoe': ap.get('dm', ''),
            'dislipidemia': ap.get('dislipidemia', ''),
            'escolaridade': ap.get('escolaridade', ''),
            'cor_raca': ap.get('cor_raca', ''),
            'age_csv': laudo.get('age', ''),
            'sex_csv': laudo.get('sex', ''),
            'has_csv': laudo.get('has_db', ''),
            'dm_csv': laudo.get('dm_db', ''),
            'od_quality': laudo['od_quality'],
            'oe_quality': laudo['oe_quality'],
            'url_od': laudo['url_od'],
            'url_oe': laudo['url_oe'],
            'local_od': local_od,
            'local_oe': local_oe,
            'diagnosis': laudo.get('diagnosis', ''),
        })

    print(f"\n--- RESUMO ---")
    print(f"Pacientes matched com genótipo: {len(matches)}")
    print(f"Pacientes com pelo menos 1 imagem usável: {patients_with_images}")
    print(f"Total de imagens usáveis: {total_images}")

    # Genotype distribution
    from collections import Counter
    genotypes = Counter(r['genotype'] for r in results if r['local_od'] or r['local_oe'])
    print(f"\nDistribuição por genótipo (com imagens):")
    for g in ['e2e2', 'e2e3', 'e2e4', 'e3e3', 'e3e4', 'e4e4']:
        n = genotypes.get(g, 0)
        print(f"  {g}: {n}")

    # Show unmatched
    if unmatched:
        print(f"\n--- UNMATCHED CSV ({len(unmatched)}) ---")
        for laudo, best_ap, best_score in sorted(unmatched, key=lambda x: -x[2]):
            best_info = f" (closest: {best_ap['name_normalized']}, score={best_score:.2f})" if best_ap else ""
            print(f"  {laudo['name_normalized']}{best_info}")

    # Write output
    if execute:
        out_path = SCRIPT_DIR / 'matched_patients.csv'
        fieldnames = list(results[0].keys()) if results else []
        with open(out_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(results)
        print(f"\n>>> Gravado: {out_path} ({len(results)} registros)")
    else:
        print(f"\n[PREVIEW] Use --execute para gerar matched_patients.csv")


if __name__ == '__main__':
    main()
